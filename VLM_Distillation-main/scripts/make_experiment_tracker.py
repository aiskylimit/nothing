"""Generate `experiments/experiment_tracker.xlsx` — the live tracker for the
Phase-0 paper sprint.

Re-run this script whenever you want to (a) regenerate a fresh tracker from
scratch, (b) add planned runs, or (c) refresh the structure after editing the
schema below. Existing edited tracker data is NOT preserved — copy any manual
edits out before re-running, or rename the output path.

Six sheets:
  README       — what each sheet is for + workflow notes.
  Runs         — one row per training run. Edit `status`, `wall_clock_h`, etc.
  Results      — long-format (run × benchmark) score table. Edit `score` cells.
  Main_Table   — pivoted view for the paper. Auto-populated via VLOOKUP from Results.
  UMI          — per-pair Unit-Mismatch Index numbers (filled from scripts/compute_umi.py output).
  Compute_Log  — day-by-day compute vs the sprint plan.

Usage:
  /usr/bin/python3 scripts/make_experiment_tracker.py
  # writes experiments/experiment_tracker.xlsx (overwrites)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
PLANNED_FILL = PatternFill("solid", fgColor="FFF2CC")
ABLATION_FILL = PatternFill("solid", fgColor="F4CCCC")

PROJECT_PAIR = "qwen25_to_qwen2"
TEACHER = "Qwen/Qwen2.5-VL-7B-Instruct"
STUDENT = "Qwen/Qwen2-VL-2B-Instruct"

# Planned training runs for the 7-day sprint.
RUNS = [
    {
        "run_id": "ce_only_run1",
        "method": "CE-only",
        "kd_loss_type": "ce_only",
        "category": "baseline",
        "extra_args": "",
        "script": "script_train/ce_only/train_qwen25_teacher_7b_qwen2_student_2b_ce_only.sh",
    },
    {
        "run_id": "emkd_run1",
        "method": "EM-KD-only",
        "kd_loss_type": "emkd",
        "category": "single-side",
        "extra_args": "em_kd_alpha=0.5, beta=0.25, gamma=25.0, T=1.0",
        "script": "script_train/em_kd/train_qwen25_teacher_7b_qwen2_student_2b_emkd.sh",
    },
    {
        "run_id": "sre_run1",
        "method": "SRE-only",
        "kd_loss_type": "sre",
        "category": "single-side",
        "extra_args": "layer_mapping=[27], sre_use_projector=True, sre_geom_loss_weight=50",
        "script": "script_train/sre/train_qwen25_teacher_7b_qwen2_student_2b_sre.sh",
    },
    {
        "run_id": "unit_aligned_run1",
        "method": "Unit-Aligned",
        "kd_loss_type": "unit_aligned",
        "category": "joint",
        "extra_args": "layer_mapping=[27], joint_ce=0.5, joint_emkd=1.0, joint_sre=1.0",
        "script": "script_train/unit_aligned/train_qwen25_teacher_7b_qwen2_student_2b_unit_aligned.sh",
    },
    {
        "run_id": "unit_aligned_noproj_ablation",
        "method": "Unit-Aligned (no projector)",
        "kd_loss_type": "unit_aligned",
        "category": "ablation",
        "extra_args": "sre_use_projector=False (fallback to crop)",
        "script": "script_train/unit_aligned/train_..._unit_aligned.sh + --sre_use_projector false",
    },
]

# Shared run config (current sprint defaults).
SHARED_CONFIG = {
    "teacher": TEACHER,
    "student": STUDENT,
    "lora_r": 64,
    "lora_alpha": 64,
    "lr": "1e-5",
    "per_device_bs": 1,
    "grad_accum": 8,
    "epochs": 1,
    "percent_data": 0.15,
    "seed": 1337,
    "max_len": 2048,
    "image_resolution": "mid",
    "bf16": True,
}

# Eval suites: (name, list of dataset short labels in the order they appear in
# configs/eval/<suite>.json — keep in sync if you edit those configs).
FAST_SIGNAL = ["MMBench_DEV_EN", "MMStar", "MMMU_DEV_VAL", "MathVista_MINI"]
MAIN_TABLE_EXTRA = ["MMVet", "AI2D_TEST_NO_MASK", "ChartQA_TEST", "POPE"]
FULL_EXTRA = ["RealWorldQA", "BLINK", "OCRBench", "DocVQA_VAL"]

SUITES = {
    "fast_signal": FAST_SIGNAL,
    "main_table": FAST_SIGNAL + MAIN_TABLE_EXTRA,
    "full": FAST_SIGNAL + MAIN_TABLE_EXTRA + FULL_EXTRA,
}


def _style_header(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_readme(ws) -> None:
    ws.title = "README"
    rows = [
        ["VLM Distillation — Phase-0 Paper Sprint Tracker", ""],
        ["", ""],
        ["Submission target", "ARR May cycle, internal cutoff 2026-05-22"],
        ["Pair", f"{TEACHER} -> {STUDENT}"],
        ["Branch", "phase-0-paper-sprint"],
        ["Plan", "phase_0_sprint_plan.md"],
        ["", ""],
        ["Sheet", "What it's for"],
        ["Runs", "One row per training run. Edit status/wall_clock/notes as runs complete. Yellow = planned, Red = ablation."],
        ["Results", "Long-format (run x benchmark) score table. Fill the `score` column after each VLMEvalKit run."],
        ["Main_Table", "Pivoted view for the paper. Uses VLOOKUP into Results — DON'T edit cells, edit Results."],
        ["UMI", "Per-pair Unit-Mismatch Index from scripts/compute_umi.py output. Copy the JSON fields into the row."],
        ["Compute_Log", "Daily compute & API spend vs the sprint plan. Track to avoid running out of budget mid-week."],
        ["", ""],
        ["Workflow", ""],
        ["1", "Edit `Runs` to mark a run as in-progress / done."],
        ["2", "After eval finishes, fill the four/eight/twelve `score` rows for that run in `Results`."],
        ["3", "`Main_Table` updates automatically. Decide go/no-go at end of Day 6 using `Avg_fast` column."],
        ["4", "Add new runs by appending to `Runs` AND duplicating its result rows in `Results`."],
        ["5", "If you change the schema, re-run `python3 scripts/make_experiment_tracker.py` (note: overwrites)."],
        ["", ""],
        ["Decision criterion (from phase_0_sprint_plan.md)", ""],
        ["Submit", "Unit-Aligned >= max(EM-KD-only, SRE-only) - 0.3 pts averaged AND > CE-only avg"],
        ["Reframe", "Unit-Aligned < CE-only -> rewrite as robustness/negative-result paper for a workshop"],
        ["Abandon sprint", "All KD methods < CE-only by >= 1 pt -> add Qwen3-VL pair, target ARR June"],
    ]
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            elif r == 8 or r == 15 or r == 23:
                cell.fill = SECTION_FILL
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _set_col_widths(ws, [38, 78])


def build_runs(ws) -> None:
    ws.title = "Runs"
    headers = [
        "run_id", "method", "category", "kd_loss_type", "extra_args", "script",
        "teacher", "student", "lora_r", "lora_alpha", "lr", "per_device_bs",
        "grad_accum", "epochs", "percent_data", "seed", "max_len",
        "status", "start_date", "end_date", "wall_clock_h", "ckpt_path", "final_step", "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for r, run in enumerate(RUNS, start=2):
        values = [
            run["run_id"], run["method"], run["category"], run["kd_loss_type"],
            run["extra_args"], run["script"],
            SHARED_CONFIG["teacher"], SHARED_CONFIG["student"],
            SHARED_CONFIG["lora_r"], SHARED_CONFIG["lora_alpha"],
            SHARED_CONFIG["lr"], SHARED_CONFIG["per_device_bs"],
            SHARED_CONFIG["grad_accum"], SHARED_CONFIG["epochs"],
            SHARED_CONFIG["percent_data"], SHARED_CONFIG["seed"], SHARED_CONFIG["max_len"],
            "planned", "", "", "", "", "", "",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if run["category"] == "ablation":
                cell.fill = ABLATION_FILL
            else:
                cell.fill = PLANNED_FILL

    widths = [
        22, 22, 12, 16, 42, 60, 26, 26, 8, 10, 8, 14, 12, 8, 14, 6, 9,
        12, 12, 12, 14, 38, 10, 40,
    ]
    _set_col_widths(ws, widths)


def build_results(ws) -> None:
    ws.title = "Results"
    headers = [
        "run_id", "method", "suite", "dataset", "score", "n_samples", "judge_model",
        "eval_hours", "judge_usd", "eval_date", "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    # For the sprint, pre-populate fast_signal (4 datasets) for every run.
    # User adds main_table / full rows manually when expanding.
    row = 2
    for run in RUNS:
        for dataset in SUITES["fast_signal"]:
            ws.cell(row=row, column=1, value=run["run_id"])
            ws.cell(row=row, column=2, value=run["method"])
            ws.cell(row=row, column=3, value="fast_signal")
            ws.cell(row=row, column=4, value=dataset)
            # score, n_samples, judge_model, eval_hours, judge_usd, eval_date, notes left blank
            row += 1
    # And the same 4 runs × 8 main_table datasets, so the main table pivot
    # has cells to look up even before they're filled.
    for run in RUNS:
        if run["category"] == "ablation":
            continue  # skip ablation for main table; only run on fast_signal for headline
        for dataset in MAIN_TABLE_EXTRA:
            ws.cell(row=row, column=1, value=run["run_id"])
            ws.cell(row=row, column=2, value=run["method"])
            ws.cell(row=row, column=3, value="main_table")
            ws.cell(row=row, column=4, value=dataset)
            row += 1

    _set_col_widths(ws, [22, 22, 14, 22, 10, 12, 26, 12, 12, 12, 40])


def build_main_table(ws) -> None:
    ws.title = "Main_Table"
    methods = [r for r in RUNS if r["category"] != "ablation"]
    datasets = SUITES["main_table"]

    # Row 1: title.
    ws.cell(row=1, column=1, value="Main Paper Table (Method x Benchmark)").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value="Edit Results sheet; this table refreshes via VLOOKUP. Blank cells mean no result yet.").font = Font(italic=True, color="666666")

    # Header.
    header_row = 4
    ws.cell(row=header_row, column=1, value="Method")
    for c, ds in enumerate(datasets, start=2):
        ws.cell(row=header_row, column=c, value=ds)
    ws.cell(row=header_row, column=len(datasets) + 2, value="Avg_fast (4)")
    ws.cell(row=header_row, column=len(datasets) + 3, value="Avg_main (8)")
    _style_header(ws, header_row, len(datasets) + 3)

    for r, run in enumerate(methods, start=header_row + 1):
        ws.cell(row=r, column=1, value=run["method"])
        for c, ds in enumerate(datasets, start=2):
            # IFERROR(VLOOKUP into Results filtered by run_id+dataset). We use
            # SUMPRODUCT instead of array formulas for openpyxl portability.
            formula = (
                f"=IFERROR("
                f"INDEX(Results!$E:$E, "
                f"MATCH(1, (Results!$A:$A=\"{run['run_id']}\")*(Results!$D:$D=\"{ds}\"), 0))"
                f", \"\")"
            )
            ws.cell(row=r, column=c, value=formula)
        # Avg_fast = average over first 4 datasets (fast_signal).
        first_col = get_column_letter(2)
        last_fast_col = get_column_letter(1 + len(FAST_SIGNAL))
        ws.cell(
            row=r, column=len(datasets) + 2,
            value=f"=IFERROR(AVERAGE({first_col}{r}:{last_fast_col}{r}), \"\")",
        )
        # Avg_main = average over all 8 main_table datasets.
        last_main_col = get_column_letter(1 + len(datasets))
        ws.cell(
            row=r, column=len(datasets) + 3,
            value=f"=IFERROR(AVERAGE({first_col}{r}:{last_main_col}{r}), \"\")",
        )

    # Decision-row helper: Unit-Aligned - max(EM-KD, SRE), expected delta.
    decision_row = header_row + len(methods) + 2
    ws.cell(row=decision_row, column=1, value="Delta (Unit-Aligned - max(EM-KD, SRE)) — submit if >= -0.3").font = Font(bold=True)
    ua_row = header_row + 1 + next(i for i, m in enumerate(methods) if m["method"] == "Unit-Aligned")
    em_row = header_row + 1 + next(i for i, m in enumerate(methods) if m["method"] == "EM-KD-only")
    sre_row = header_row + 1 + next(i for i, m in enumerate(methods) if m["method"] == "SRE-only")
    for c in range(2, len(datasets) + 4):
        col = get_column_letter(c)
        ws.cell(
            row=decision_row, column=c,
            value=f"=IFERROR({col}{ua_row} - MAX({col}{em_row}, {col}{sre_row}), \"\")",
        )

    widths = [26] + [16] * len(datasets) + [14, 14]
    _set_col_widths(ws, widths)


def build_umi(ws) -> None:
    ws.title = "UMI"
    headers = [
        "pair_name", "teacher", "student", "vocab_jaccard_on_corpus",
        "student_vocab_used", "teacher_vocab_used", "vision_token_ratio_mean",
        "umi_text_component", "umi_vision_component", "umi",
        "response_sample", "image_sample", "seed", "computed_date", "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    ws.cell(row=2, column=1, value=PROJECT_PAIR)
    ws.cell(row=2, column=2, value=TEACHER)
    ws.cell(row=2, column=3, value=STUDENT)
    # Other columns left blank — fill from `python scripts/compute_umi.py ... --out_path artifacts/umi.json`
    ws.cell(row=4, column=1, value="(future pairs for ARR June: Qwen3-VL-8B->Qwen2.5-VL-3B, LLaVA-OV-7B->Qwen2.5-VL-3B)").font = Font(italic=True, color="666666")

    _set_col_widths(ws, [22, 26, 26, 18, 16, 16, 18, 16, 16, 10, 14, 12, 6, 14, 36])


def build_compute_log(ws) -> None:
    ws.title = "Compute_Log"
    headers = [
        "date", "day_num", "planned_task", "actual_task",
        "gpu_hours_used_today", "gpu_hours_cumulative", "judge_usd_today", "judge_usd_cumulative",
        "blocker", "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    plan_rows = [
        ("2026-05-15", 1, "Code fixes + UMI computation + paper skeleton stub"),
        ("2026-05-16", 2, "Teacher logit cache + CE-only training launch + paper skeleton"),
        ("2026-05-17", 3, "EM-KD-only training launch + VLMEvalKit setup"),
        ("2026-05-18", 4, "SRE-only training launch + CE-only fast_signal eval"),
        ("2026-05-19", 5, "Unit-Aligned training launch + EM-KD-only fast_signal eval"),
        ("2026-05-20", 6, "Unit-Aligned eval + SRE-only eval + ablation run"),
        ("2026-05-21", 7, "Tables + write results/discussion + camera-ready polish"),
        ("2026-05-22", 8, "Final reproducibility check + SUBMIT"),
    ]
    for r, (date, day, task) in enumerate(plan_rows, start=2):
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=day)
        ws.cell(row=r, column=3, value=task)

    # Cumulative formulas (auto-sum).
    last_data_row = 1 + len(plan_rows)
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=6, value=f"=SUM($E$2:E{r})")
        ws.cell(row=r, column=8, value=f"=SUM($G$2:G{r})")

    _set_col_widths(ws, [12, 8, 56, 56, 16, 18, 14, 18, 26, 36])


def main() -> None:
    project_dir = Path(__file__).resolve().parents[1]
    out_path = project_dir / "experiments" / "experiment_tracker.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_readme(wb.active)
    build_runs(wb.create_sheet("Runs"))
    build_results(wb.create_sheet("Results"))
    build_main_table(wb.create_sheet("Main_Table"))
    build_umi(wb.create_sheet("UMI"))
    build_compute_log(wb.create_sheet("Compute_Log"))

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
